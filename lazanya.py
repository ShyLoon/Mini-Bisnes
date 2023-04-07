import os
import xlsxwriter
from openpyxl import load_workbook
import openpyxl

print("hello")
def createFile():  # создание файла
    
    workbook = xlsxwriter.Workbook("test.xlsx")
    workbook.close()
    #worksheet = workbook.add_worksheet()


def add_user(login: str, password: str) -> bool:

    fn = 'test.xlsx'
    wb = load_workbook(fn)
    ws = wb['sheet1']
    ws.append([f'{login}',f'{password}'])
    wb.save(fn)
    wb.close()
    # with open('users.xlsx', 'a') as worksheet:
    #     worksheet.write(f'{login}:{password}\n')  
    # return True


def search_user(login: str, password: str) -> bool:
    book = openpyxl.open("users.xml", read_only=True)
    with open('users.xlsx', 'r') as search:
        users = search.read().splitlines()  # ищем пользователя во всём файле 

    for user in users:
        args = user.split(':')
        if login == args[0] and password == args[1]:  # Если пользователь с таким логином и паролем существует
            return True
    return False


def reg():
    

    createFile()
    

    while True:
        print('''Добро пожаловать! Выберите пункт меню:
        1. Вход
        2. Регистрация
        3. Выход''')

        user_input = input()
        if user_input == '1': 
            print('Введите логин:')
            login = input()

            print('Введите пароль:')
            password = input()

            result = search_user(login, password)

            if result:
                print('Вы вошли в систему')
                break  # Выходим из цикла
            else:
                print('Неверный логин или пароль')

        elif user_input == '2':
            print('Введите логин:')
            login = input()

            print('Введите пароль:')
            password = input()

            print('Повторите пароль:')
            password_repeat = input()

            if password != password_repeat:
                print('Пароли не совпадают!')
                continue

            result = add_user(login, password) 

            if not result:
                print('Пользователь с таким логином уже существует')
            else:
                print('Регистрация прошла успешно!')

        elif user_input == '3':
            print('Не буду я у вас кушать')
            break  


reg()

# def dish():
#     check_list = [""]
#     summa = 0
#     dish = 0
#     while True:        
#         print('''Хелоу,  
#         Что вы хотите?
#         1. Собрать лозанью
#         2. Узнать стоимость
#         3. Покинуть заведение
#         4. Получить чек''')


#         user_control = input()
#         if user_control == '1':
#             dish += 1
#             i = 0
#             for i in range(5):
#                 print('''Чтобы заказть блюдо, вам нужно выбрать ингридиенты:
#                 1. Фарш - 800 грамм 
#                 2. Помидоры
#                 3. Лук
#                 4. Листы лазаньи
#                 5. Твёрдый сыр ''')

#                 collection = input()
#                 match collection:
#                     case '1':
#                         check_list.append('farsh')
#                         summa += 2300
#                     case '2':
#                         check_list.append('tomato')
#                         summa += 500
#                     case '3':
#                         check_list.append('lyk')
#                         summa += 250
#                     case '4':
#                         check_list.append('listLazan')
#                         summa += 780
#                     case '5':
#                         check_list.append('cheese')
#                         summa += 350
#             i += 1
#             continue
#         elif (user_control == '2'):
#             print('Стоимость вашего блюда:', summa)
#             break


# dish()

            



            
            
